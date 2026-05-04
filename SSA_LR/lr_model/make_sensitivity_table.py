#!/usr/bin/env python3
"""
make_sensitivity_table.py — Reproduce TR2025 Table VI.D1.

Runs the full model pipeline for ultimate TFR = 1.6, 1.9, and 2.1
(all with intermediate economic/mortality assumptions) and writes the
sensitivity table to Excel.

Usage:
    cd SSA_LR/lr_model
    python make_sensitivity_table.py
"""

import sys
import pathlib
import numpy as np
import pandas as pd

ROOT   = pathlib.Path(__file__).resolve().parent
PY_DIR = ROOT / "python"
sys.path.insert(0, str(PY_DIR))

import assumptions as _asm
import economics   as _ec
import trust_fund  as _tf_mod
from demography    import run_demography
from economics     import run_economics
from beneficiaries import run_beneficiaries
from trust_fund    import run_trust_fund

# ── constants ──────────────────────────────────────────────────────────────────
TF_START   = _tf_mod.OASI_ASSETS_2024 + _tf_mod.DI_ASSETS_2024   # $2 712.1 B
PROJ_START = 2025
TARGET_AB  = -3.82   # intermediate 75-year actuarial balance target

# SSA LR Model exposure factors (§4.3)
EXP_CONTRB  = 0.518
EXP_TAXBEN  = 0.625
EXP_ADM     = 0.5
EXP_RR      = 7.0 / 12.0
EXP_PAYROLL = 0.5

PERIODS = [
    ("25-year: 2025-49", 2049),
    ("50-year: 2025-74", 2074),
    ("75-year: 2025-99", 2099),
]

TFR_NODE_MAP = {
    1.6: {2025: 1.62, 2050: 1.60, 2100: 1.60},
    1.9: {2025: 1.62, 2050: 1.90, 2100: 1.90},
    2.1: {2025: 1.62, 2050: 2.10, 2100: 2.10},
}

# ── helpers ────────────────────────────────────────────────────────────────────

def _copy_intermediate_to_custom():
    """Populate assumptions["custom"] from assumptions["intermediate"]."""
    for key in ["AWI_NOMINAL_PCT", "CPI_PCT", "EARNINGS_PCT_COMPENSATION",
                "LPR_NET", "TUP_NET"]:
        getattr(_asm, key)["custom"] = getattr(_asm, key)["intermediate"]
    for key in ["MORTALITY_ULT_REDUCTION", "MORTALITY_ULTIMATE_YEAR",
                "UNEMPLOYMENT_ULT", "REAL_INTEREST_RATE_ULT",
                "TOB_FACTOR_OASI", "TOB_FACTOR_DI", "NEW_ISSUE_YIELD"]:
        getattr(_asm, key)["custom"] = getattr(_asm, key)["intermediate"]
    _ec.CWR_DRIFT["custom"]  = _ec.CWR_DRIFT["intermediate"]
    _ec.UNEMP_PATH["custom"] = _ec.UNEMP_PATH["intermediate"]


def run_pipeline(tfr_ult):
    """Run the full four-module pipeline for the given ultimate TFR."""
    if tfr_ult == 1.9:
        scenario = "intermediate"
        _asm.SCENARIO = scenario
    else:
        scenario = "custom"
        _asm.SCENARIO = scenario
        _asm.TFR_NODES["custom"] = TFR_NODE_MAP[tfr_ult]
        _copy_intermediate_to_custom()

    demo = run_demography(scenario=scenario,  calibrate_flag=False)
    econ = run_economics(demo, scenario=scenario,  calibrate_flag=False)
    bene = run_beneficiaries(demo, econ, scenario=scenario, calibrate_flag=False)
    tf   = run_trust_fund(demo, econ, bene, scenario=scenario, calibrate_flag=False)
    return tf["tf"], tf["exhaustion"], scenario


def _irate_v(tf_df, years, scenario):
    """Return (irate dict, cumulative-discount-factor dict) for the given years."""
    irate, v = {}, {}
    cv = 1.0
    for yr in years:
        beg = (TF_START if yr == PROJ_START
               else float(tf_df.loc[yr - 1, "oasdi_assets_bn"]))
        end = float(tf_df.loc[yr, "oasdi_assets_bn"])
        avg = (beg + end) / 2.0
        tot_int = float(tf_df.loc[yr, "int_oasi_bn"] + tf_df.loc[yr, "int_di_bn"])
        irate[yr] = (tot_int / avg if avg > 10.0
                     else _asm.get_yield(yr, scenario))
        cv /= (1.0 + irate[yr])
        v[yr] = cv
    return irate, v


def _target_fund_pv(tf_df, yr2, yr2m, v):
    """PV of 1-year's cost projected to the year after period end."""
    def _c(yr):
        return float(
            tf_df.loc[yr, "oasi_bene_bn"] + tf_df.loc[yr, "di_bene_bn"] +
            tf_df.loc[yr, "admin_oasi_bn"] + tf_df.loc[yr, "admin_di_bn"] -
            tf_df.loc[yr, "rrb_oasi_bn"]  - tf_df.loc[yr, "rrb_di_bn"]
        )
    tc_end, tc_prev = _c(yr2), _c(yr2m)
    tc_next = tc_end ** 2 / tc_prev   # geometric extrapolation
    return tc_next * v[yr2]


def summarized_rates(tf_df, period_end, ben_exp, scenario):
    """
    Compute (summarized_income_rate, summarized_cost_rate, actuarial_balance)
    for a valuation window [PROJ_START, period_end].
    """
    years  = list(range(PROJ_START, period_end + 1))
    irate, v = _irate_v(tf_df, years, scenario)
    pv_targ  = _target_fund_pv(tf_df, years[-1], years[-2], v)

    S = {k: 0.0 for k in ("contrb", "taxben", "ben", "adm", "rr", "pay")}
    for yr in years:
        ir, vt = irate[yr], v[yr]
        contrb = float(tf_df.loc[yr, "oasi_tax_bn"]   + tf_df.loc[yr, "di_tax_bn"])
        taxben = float(tf_df.loc[yr, "tob_oasi_bn"]   + tf_df.loc[yr, "tob_di_bn"])
        ben    = float(tf_df.loc[yr, "oasi_bene_bn"]  + tf_df.loc[yr, "di_bene_bn"])
        adm    = float(tf_df.loc[yr, "admin_oasi_bn"] + tf_df.loc[yr, "admin_di_bn"])
        rr     = float(-(tf_df.loc[yr, "rrb_oasi_bn"] + tf_df.loc[yr, "rrb_di_bn"]))
        pay    = float(tf_df.loc[yr, "payroll_base_bn"])
        S["contrb"] += (1 + EXP_CONTRB  * ir) * contrb * vt
        S["taxben"] += (1 + EXP_TAXBEN  * ir) * taxben * vt
        S["ben"]    += (1 + ben_exp      * ir) * ben    * vt
        S["adm"]    += (1 + EXP_ADM     * ir) * adm    * vt
        S["rr"]     += (1 + EXP_RR      * ir) * rr     * vt
        S["pay"]    += (1 + EXP_PAYROLL * ir) * pay    * vt

    sir = (TF_START + S["contrb"] + S["taxben"]) / S["pay"] * 100
    scr = (S["ben"] + S["adm"] + S["rr"] + pv_targ) / S["pay"] * 100
    return round(sir, 2), round(scr, 2), round(sir - scr, 2)


def calibrate_ben_exp(tf_df, scenario):
    """Bisect ben_exp so 75-year intermediate actuarial balance = TARGET_AB."""
    lo, hi = 0.0, 3.0
    for _ in range(80):
        m  = (lo + hi) / 2.0
        ab = summarized_rates(tf_df, 2099, m, scenario)[2]
        if ab > TARGET_AB:
            lo = m
        else:
            hi = m
    return (lo + hi) / 2.0


# ── run three scenarios ────────────────────────────────────────────────────────

print("=" * 60)
print("  SSA Fertility Sensitivity Table — Table VI.D1")
print("=" * 60)

print("\n[1/3] Running TFR = 1.9 (intermediate, calibrating ben_exp)...")
tf_19, exh_19, sc_19 = run_pipeline(1.9)
ben_exp = calibrate_ben_exp(tf_19, sc_19)
print(f"      Calibrated ben_exp = {ben_exp:.6f}")
# cache so that re-runs of main.py stay consistent
(PY_DIR / "_intermediate_ben_exp.txt").write_text(f"{ben_exp:.8f}")

print("\n[2/3] Running TFR = 1.6...")
tf_16, exh_16, sc_16 = run_pipeline(1.6)

print("\n[3/3] Running TFR = 2.1...")
tf_21, exh_21, sc_21 = run_pipeline(2.1)

# ── compute table metrics ─────────────────────────────────────────────────────

runs = [
    (1.6, tf_16, exh_16, sc_16),
    (1.9, tf_19, exh_19, sc_19),
    (2.1, tf_21, exh_21, sc_21),
]

table = {}
for tfr, tf_df, exh, sc in runs:
    inc, cst, ab = [], [], []
    for _, period_end in PERIODS:
        i, c, a = summarized_rates(tf_df, period_end, ben_exp, sc)
        inc.append(i); cst.append(c); ab.append(a)
    ann_bal_2099 = round(float(tf_df.loc[2099, "oasdi_net_rate"]), 2)
    exh_year = exh["oasi"]   # SSA table uses OASI fund depletion (the binding constraint)
    table[tfr] = {"inc": inc, "cst": cst, "ab": ab,
                  "ann_2099": ann_bal_2099, "exh": exh_year}

print("\n  Results preview:")
print(f"  {'Metric':<30}  {'TFR=1.6':>8}  {'TFR=1.9':>8}  {'TFR=2.1':>8}")
for i, (lbl, _) in enumerate(PERIODS):
    print(f"  Summarized income {lbl:<15}"
          f"  {table[1.6]['inc'][i]:>8.2f}"
          f"  {table[1.9]['inc'][i]:>8.2f}"
          f"  {table[2.1]['inc'][i]:>8.2f}")
for i, (lbl, _) in enumerate(PERIODS):
    print(f"  Summarized cost   {lbl:<15}"
          f"  {table[1.6]['cst'][i]:>8.2f}"
          f"  {table[1.9]['cst'][i]:>8.2f}"
          f"  {table[2.1]['cst'][i]:>8.2f}")
for i, (lbl, _) in enumerate(PERIODS):
    print(f"  Actuarial balance {lbl:<15}"
          f"  {table[1.6]['ab'][i]:>8.2f}"
          f"  {table[1.9]['ab'][i]:>8.2f}"
          f"  {table[2.1]['ab'][i]:>8.2f}")
print(f"  Ann. balance 2099              "
      f"  {table[1.6]['ann_2099']:>8.2f}"
      f"  {table[1.9]['ann_2099']:>8.2f}"
      f"  {table[2.1]['ann_2099']:>8.2f}")
print(f"  Trust fund depletion           "
      f"  {str(table[1.6]['exh']):>8}"
      f"  {str(table[1.9]['exh']):>8}"
      f"  {str(table[2.1]['exh']):>8}")

# ── write Excel ───────────────────────────────────────────────────────────────

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter
except ImportError:
    sys.exit("openpyxl not found — run:  pip install openpyxl")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Table VI.D1"

# column widths
ws.column_dimensions["A"].width = 38
for col_letter in ["B", "C", "D"]:
    ws.column_dimensions[col_letter].width = 11

thin  = Side(style="thin")
med   = Side(style="medium")
no    = Side(style=None)

TNR   = "Times New Roman"

def _border(top=None, bottom=None, left=None, right=None):
    return Border(top=top or no, bottom=bottom or no,
                  left=left  or no, right=right or no)


def _set(ws, r, c, value=None, bold=False, italic=False,
         halign="center", valign="center", indent=0,
         num_fmt=None, top=None, bottom=None, left=None, right=None,
         size=10, font_color="000000"):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font      = Font(name=TNR, bold=bold, italic=italic, size=size,
                          color=font_color)
    cell.alignment = Alignment(horizontal=halign, vertical=valign,
                               indent=indent, wrap_text=True)
    if num_fmt:
        cell.number_format = num_fmt
    cell.border = _border(top=top, bottom=bottom, left=left, right=right)
    return cell

# ── Row 1: Title ──────────────────────────────────────────────────────────────
ws.merge_cells("A1:D1")
_set(ws, 1, 1,
     "Table VI.D1.—Sensitivity of OASDI Measures to Fertility Assumptions",
     bold=True, halign="center", size=10)
ws.row_dimensions[1].height = 16

# ── Row 2: Subtitle ───────────────────────────────────────────────────────────
ws.merge_cells("A2:D2")
_set(ws, 2, 1, "[As a percentage of taxable payroll]",
     italic=True, halign="center", size=10)
ws.row_dimensions[2].height = 14

# ── Row 3: blank ──────────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 6

# ── Row 4: column headers ─────────────────────────────────────────────────────
ws.merge_cells("A4:A5")
_set(ws, 4, 1, "Valuation period",
     halign="center", valign="bottom",
     bottom=thin, top=med)

ws.merge_cells("B4:D4")
_set(ws, 4, 2, "Ultimate total fertility rate ᵃ ᵇ",
     halign="center", bottom=thin, top=med)

# ── Row 5: TFR sub-headers ────────────────────────────────────────────────────
for j, tfr_val in enumerate([1.6, 1.9, 2.1]):
    _set(ws, 5, j + 2, tfr_val,
         halign="center", num_fmt="0.0", bottom=thin)

# ── helper: data row ──────────────────────────────────────────────────────────
TFR_COLS = [1.6, 1.9, 2.1]

def _section_header(r, label):
    _set(ws, r, 1, label, bold=True, halign="left")

def _data_row(r, label, values, num_fmt="0.00"):
    _set(ws, r, 1, label, halign="left", indent=2)
    for j, v in enumerate(values):
        _set(ws, r, j + 2, v, halign="right", num_fmt=num_fmt)

# ── Rows 6+: data ─────────────────────────────────────────────────────────────
r = 6

_section_header(r, "Summarized income rate:"); r += 1
for i, (lbl, _) in enumerate(PERIODS):
    _data_row(r, lbl, [table[t]["inc"][i] for t in TFR_COLS]); r += 1

_section_header(r, "Summarized cost rate:"); r += 1
for i, (lbl, _) in enumerate(PERIODS):
    _data_row(r, lbl, [table[t]["cst"][i] for t in TFR_COLS]); r += 1

_section_header(r, "Actuarial balance:"); r += 1
for i, (lbl, _) in enumerate(PERIODS):
    _data_row(r, lbl, [table[t]["ab"][i] for t in TFR_COLS]); r += 1

# Annual balance 2099
_set(ws, r, 1, "Annual balance for 2099", bold=True, halign="left")
for j, tfr_val in enumerate(TFR_COLS):
    _set(ws, r, j + 2, table[tfr_val]["ann_2099"],
         halign="right", num_fmt="0.00")
r += 1

# Year of depletion
_set(ws, r, 1, "Year of combined trust fund reserve depletion",
     bold=True, halign="left")
for j, tfr_val in enumerate(TFR_COLS):
    exh_val = table[tfr_val]["exh"]
    cell_val = int(exh_val) if isinstance(exh_val, (int, float)) else exh_val
    _set(ws, r, j + 2, cell_val, halign="right")
last_data_row = r
r += 1

# bottom border on last data row
for c in range(1, 5):
    ws.cell(row=last_data_row, column=c).border = _border(bottom=med)

# ── Footnotes ─────────────────────────────────────────────────────────────────
r += 1
ws.merge_cells(f"A{r}:D{r}")
_set(ws, r, 1,
     "ᵃ Intermediate assumptions for fertility, mortality, immigration, "
     "and economic variables from the 2025 Trustees Report are used as the baseline.",
     size=9, halign="left", valign="top")
ws.row_dimensions[r].height = 28
r += 1
ws.merge_cells(f"A{r}:D{r}")
_set(ws, r, 1,
     "ᵇ The ultimate total fertility rate is the assumed long-run value, reached by 2050.",
     size=9, halign="left")

# ── Sheet 2: Model vs SSA comparison ─────────────────────────────────────────

# SSA reference values from TR2025 Table VI.D1 screenshot
SSA = {
    "inc": [[14.24, 14.24, 14.24],   # 25-yr
            [13.91, 13.88, 13.86],   # 50-yr
            [13.85, 13.79, 13.75]],  # 75-yr
    "cst": [[16.89, 16.91, 16.92],
            [17.56, 17.28, 17.10],
            [18.34, 17.61, 17.15]],
    "ab":  [[-2.65, -2.67, -2.67],
            [-3.65, -3.40, -3.24],
            [-4.49, -3.82, -3.40]],
    "ann_2099": [-7.39, -4.84, -3.46],
    "exh":      [2034,  2034,  2034],
}

wc = wb.create_sheet("Model vs SSA")

# column widths
wc.column_dimensions["A"].width = 32
wc.column_dimensions["B"].width = 8
for col_letter in "CDEFGHIJKLMNO":
    wc.column_dimensions[col_letter].width = 8

RED  = "C00000"
GRN  = "375623"
GREY = "595959"

def _sc(ws, r, c, value=None, bold=False, halign="center", num_fmt=None,
        font_color="000000", bg=None, size=9, indent=0):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = Font(name=TNR, bold=bold, size=size, color=font_color)
    cell.alignment = Alignment(horizontal=halign, vertical="center",
                               indent=indent, wrap_text=False)
    if num_fmt:
        cell.number_format = num_fmt
    if bg:
        from openpyxl.styles import PatternFill
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell

# Header
wc.merge_cells("A1:O1")
_sc(wc, 1, 1, "Table VI.D1 — Model vs SSA Reference (TR2025)",
    bold=True, halign="center", size=11)
wc.row_dimensions[1].height = 18

# Column group headers  row 2: label | [TFR 1.6: SSA / Model / Diff] | [TFR 1.9: ...] | [TFR 2.1: ...]
_sc(wc, 2, 1, "Metric", bold=True, halign="left", size=9)
for j, tfr_val in enumerate([1.6, 1.9, 2.1]):
    base_col = 2 + j * 4   # cols 2,6,10
    wc.merge_cells(start_row=2, start_column=base_col, end_row=2, end_column=base_col+3)
    _sc(wc, 2, base_col, f"TFR = {tfr_val}", bold=True, halign="center",
        bg="D9D9D9", size=9)

# Sub-headers row 3
_sc(wc, 3, 1, "", bold=True)
for j in range(3):
    base_col = 2 + j * 4
    _sc(wc, 3, base_col,     "SSA",   bold=True, halign="center", bg="F2F2F2")
    _sc(wc, 3, base_col + 1, "Model", bold=True, halign="center", bg="F2F2F2")
    _sc(wc, 3, base_col + 2, "Diff",  bold=True, halign="center", bg="F2F2F2")
    _sc(wc, 3, base_col + 3, "%Diff", bold=True, halign="center", bg="F2F2F2")

wc.row_dimensions[3].height = 14

def _cmp_rows(start_r, label, ssa_vals, model_vals, num_fmt="0.00", is_year=False):
    r = start_r
    for i, (period_lbl, _) in enumerate(PERIODS):
        _sc(wc, r, 1, f"  {period_lbl}", halign="left", size=9)
        for j in range(3):
            ssa_v   = ssa_vals[i][j]
            mod_v   = model_vals[i][j]
            diff    = round(mod_v - ssa_v, 2)
            pct     = round((mod_v - ssa_v) / abs(ssa_v) * 100, 1) if ssa_v != 0 else 0
            base_col = 2 + j * 4
            _sc(wc, r, base_col,     ssa_v, num_fmt=num_fmt)
            _sc(wc, r, base_col + 1, mod_v, num_fmt=num_fmt)
            col = RED if diff > 0.05 or diff < -0.05 else GRN
            _sc(wc, r, base_col + 2, diff,  num_fmt="+0.00;-0.00;0.00",
                font_color=col)
            _sc(wc, r, base_col + 3, pct,   num_fmt="+0.0;-0.0;0.0",
                font_color=col)
        r += 1
    return r

def _single_row(r, label, ssa_vals, model_vals, num_fmt="0.00", is_year=False):
    _sc(wc, r, 1, label, halign="left", bold=True, size=9)
    for j in range(3):
        ssa_v = ssa_vals[j]
        mod_v = model_vals[j]
        if is_year:
            diff  = (mod_v if isinstance(mod_v, int) else 9999) - ssa_v
            base_col = 2 + j * 4
            _sc(wc, r, base_col,     ssa_v, num_fmt="0")
            _sc(wc, r, base_col + 1, mod_v, num_fmt="0")
            col = GRN if diff == 0 else RED
            _sc(wc, r, base_col + 2, diff,  num_fmt="+0;-0;0", font_color=col)
            _sc(wc, r, base_col + 3, "",    font_color=col)
        else:
            diff = round(mod_v - ssa_v, 2)
            pct  = round((mod_v - ssa_v) / abs(ssa_v) * 100, 1) if ssa_v != 0 else 0
            base_col = 2 + j * 4
            _sc(wc, r, base_col,     ssa_v, num_fmt=num_fmt)
            _sc(wc, r, base_col + 1, mod_v, num_fmt=num_fmt)
            col = RED if abs(diff) > 0.05 else GRN
            _sc(wc, r, base_col + 2, diff,  num_fmt="+0.00;-0.00;0.00",
                font_color=col)
            _sc(wc, r, base_col + 3, pct,   num_fmt="+0.0;-0.0;0.0",
                font_color=col)

r = 4

# Summarized income rate
_sc(wc, r, 1, "Summarized income rate:", bold=True, halign="left"); r += 1
r = _cmp_rows(r, "inc", SSA["inc"],
              [[table[t]["inc"][i] for t in TFR_COLS] for i in range(3)])

# Summarized cost rate
_sc(wc, r, 1, "Summarized cost rate:", bold=True, halign="left"); r += 1
r = _cmp_rows(r, "cst", SSA["cst"],
              [[table[t]["cst"][i] for t in TFR_COLS] for i in range(3)])

# Actuarial balance
_sc(wc, r, 1, "Actuarial balance:", bold=True, halign="left"); r += 1
r = _cmp_rows(r, "ab", SSA["ab"],
              [[table[t]["ab"][i] for t in TFR_COLS] for i in range(3)])

# Annual balance 2099
_single_row(r, "Annual balance for 2099",
            SSA["ann_2099"], [table[t]["ann_2099"] for t in TFR_COLS]); r += 1

# Depletion year
_single_row(r, "Year of TF depletion",
            SSA["exh"],
            [table[t]["exh"] if not isinstance(table[t]["exh"], str)
             else 9999 for t in TFR_COLS],
            is_year=True); r += 1

# ── save ──────────────────────────────────────────────────────────────────────
out_path = ROOT / "table_VI_D1_fertility_sensitivity.xlsx"
wb.save(str(out_path))
print(f"\n[Done] Excel table saved to:\n  {out_path}\n")
